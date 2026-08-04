#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Genera un reporte general del sistema de pagos en Excel con tablas y graficas nativas.

Uso:
    python generar_reporte.py                          # usa DATABASE_URL del entorno o de .env
    python generar_reporte.py --salida ruta.xlsx       # ruta de salida personalizada

Requisitos:
    pip install -r requirements.txt
"""
import os
import sys
import argparse
from datetime import datetime

try:
    from dotenv import load_dotenv

    load_dotenv(os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), '.env'))
except ImportError:
    pass

try:
    import psycopg2
except ImportError:
    psycopg2 = None

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

INDIGO = "4F46E5"
LIGHT = "EEF2FF"
HEADER_FILL = PatternFill("solid", fgColor=INDIGO)
LIGHT_FILL = PatternFill("solid", fgColor=LIGHT)
HEADER_FONT = Font(bold=True, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=16, color=INDIGO)
SECTION_FONT = Font(bold=True, size=12, color=INDIGO)
MONEY_FMT = '$#,##0.00'
INT_FMT = '#,##0'
DATE_FMT = 'DD/MM/YYYY'

QUERIES = {
    "alumnos": """
        SELECT a.id, a.nombre, a.primer_apellido, a.segundo_apellido, a.email,
               a.telefono, a.grado, a.beca_id, COALESCE(b.porcentaje, 0) AS beca_porcentaje,
               a.created_at
        FROM alumnos a
        LEFT JOIN becas b ON a.beca_id = b.id
        ORDER BY a.id
    """,
    "pagos": """
        SELECT p.id,
               COALESCE(a.nombre || ' ' || a.primer_apellido || ' ' || COALESCE(a.segundo_apellido, ''), '') AS alumno,
               COALESCE(tp.concepto, '') AS concepto,
               COALESCE(tp.monto, p.monto_final) AS monto_original,
               COALESCE(p.beca_porcentaje, 0) AS beca_porcentaje,
               p.monto_final, p.semana, p.mes, p.estado, p.created_at
        FROM pagos p
        JOIN alumnos a ON p.alumno_id = a.id
        LEFT JOIN tipos_pago tp ON p.tipo_pago_id = tp.id
        ORDER BY p.id
    """,
    "comprobantes": """
        SELECT c.id, c.alumno_id,
               COALESCE(a.nombre || ' ' || a.primer_apellido || ' ' || COALESCE(a.segundo_apellido, ''), '') AS alumno,
               c.concepto, c.monto, c.metodo_pago, c.observaciones, c.created_at
        FROM comprobantes c
        LEFT JOIN alumnos a ON c.alumno_id = a.id
        ORDER BY c.id
    """,
    "inscripciones": """
        SELECT i.id, i.alumno_id,
               COALESCE(a.nombre || ' ' || a.primer_apellido || ' ' || COALESCE(a.segundo_apellido, ''), '') AS alumno,
               i.ciclo_escolar, i.grado, i.estado, i.monto_inscripcion, i.fecha_inscripcion
        FROM inscripciones i
        LEFT JOIN alumnos a ON i.alumno_id = a.id
        ORDER BY i.id
    """,
    "becas": """
        SELECT id, nombre, porcentaje, estado, descripcion
        FROM becas
        ORDER BY id
    """,
    "tipos_pago": """
        SELECT id, concepto, monto, tipo
        FROM tipos_pago
        ORDER BY id
    """,
}


def capitalizar(valor):
    s = str(valor or "").strip()
    return s[0].upper() + s[1:] if s else ""


def formatear_fecha(valor):
    if not valor:
        return None
    try:
        return valor if isinstance(valor, datetime) else datetime.fromisoformat(str(valor))
    except (ValueError, TypeError):
        return None


def _fetch(conn, query):
    with conn.cursor() as cur:
        cur.execute(query)
        cols = [d[0] for d in cur.description]
        return [dict(zip(cols, row)) for row in cur.fetchall()]


def conectar():
    if psycopg2 is None:
        sys.exit("Falta la dependencia 'psycopg2-binary'. Instala con: pip install -r requirements.txt")
    url = os.environ.get("DATABASE_URL")
    if not url:
        sys.exit("Falta la variable de entorno DATABASE_URL (usa el .env del back-end_pagos)")
    return psycopg2.connect(url, sslmode="require")


def escribir_tabla(ws, fila_inicial, encabezados, filas, anchos,
                   monedas=(), enteros=(), fechas=(), alternar=True):
    for j, enc in enumerate(encabezados, start=1):
        celda = ws.cell(row=fila_inicial, column=j, value=enc)
        celda.font = HEADER_FONT
        celda.fill = HEADER_FILL
        celda.alignment = Alignment(horizontal="center")
    for i, fila in enumerate(filas, start=fila_inicial + 1):
        for j, valor in enumerate(fila, start=1):
            celda = ws.cell(row=i, column=j, value=valor)
            idx = j - 1
            if idx in monedas:
                celda.number_format = MONEY_FMT
            elif idx in enteros:
                celda.number_format = INT_FMT
            elif idx in fechas and isinstance(valor, datetime):
                celda.number_format = DATE_FMT
            if alternar and (i - fila_inicial) % 2 == 0:
                celda.fill = LIGHT_FILL
    for j, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(j)].width = ancho
    ultima_fila = fila_inicial + len(filas)
    ws.auto_filter.ref = f"A{fila_inicial}:{get_column_letter(len(encabezados))}{ultima_fila}"
    ws.freeze_panes = f"A{fila_inicial + 1}"
    return ultima_fila


def crear_hoja_resumen(wb, data):
    ws = wb.active
    ws.title = "Resumen"

    ws["A1"] = "REPORTE GENERAL DEL SISTEMA DE PAGOS"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    ws["A2"].font = Font(italic=True, color="6B7280")

    pagos = data["pagos"]
    pagados = [p for p in pagos if p["estado"] == "pagado"]
    total_pagado = sum(float(p["monto_final"] or 0) for p in pagados)
    pendientes = len([p for p in pagos if p["estado"] == "pendiente"])
    vencidos = len([p for p in pagos if p["estado"] == "vencido"])

    kpis = [
        ("Total Pagado", total_pagado, "moneda"),
        ("Registros de Pagos", len(pagos), "entero"),
        ("Pagos Pendientes", pendientes, "entero"),
        ("Pagos Vencidos", vencidos, "entero"),
        ("Total Alumnos", len(data["alumnos"]), "entero"),
        ("Total Comprobantes", len(data["comprobantes"]), "entero"),
        ("Total Inscripciones", len(data["inscripciones"]), "entero"),
        ("Total Becas", len(data["becas"]), "entero"),
        ("Total Tipos de Pago", len(data["tipos_pago"]), "entero"),
    ]

    ws["A4"] = "Resumen General"
    ws["A4"].font = SECTION_FONT
    for i, (nombre, valor, tipo) in enumerate(kpis, start=5):
        ws.cell(row=i, column=1, value=nombre)
        celda = ws.cell(row=i, column=2, value=valor)
        celda.number_format = MONEY_FMT if tipo == "moneda" else INT_FMT

    becas_por_alumno = data.get("becas_por_alumno", [])
    fila = 15
    ws.cell(row=fila, column=1, value="Alumnos por Beca").font = SECTION_FONT
    escribir_tabla(
        ws,
        fila + 1,
        ["Beca", "Porcentaje", "Cantidad"],
        [[b["nombre"], b["porcentaje"], b["cantidad"]] for b in becas_por_alumno],
        [30, 14, 12],
        enteros={1, 2},
    )

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18


def crear_hoja_datos(wb, titulo, encabezados, filas, anchos, monedas=(), enteros=(), fechas=()):
    ws = wb.create_sheet(title=titulo)
    escribir_tabla(ws, 1, encabezados, filas, anchos, monedas, enteros, fechas)
    return ws


def crear_hoja_graficas(wb, data):
    ws = wb.create_sheet(title="Graficas")

    pagos = data["pagos"]
    ganancias_por_mes = {}
    for p in pagos:
        if p["estado"] == "pagado":
            mes = p["mes"] or "Sin mes"
            ganancias_por_mes[mes] = ganancias_por_mes.get(mes, 0) + float(p["monto_final"] or 0)
    ganancias_filas = [[m, round(v, 2)] for m, v in ganancias_por_mes.items()]
    if not ganancias_filas:
        ganancias_filas = [["Sin datos", 0]]

    pagados = len([p for p in pagos if p["estado"] == "pagado"])
    pendientes = len([p for p in pagos if p["estado"] == "pendiente"])
    vencidos = len([p for p in pagos if p["estado"] == "vencido"])
    estados_filas = [
        ["Pagados", pagados],
        ["Pendientes", pendientes],
        ["Vencidos", vencidos],
    ]

    becas_por_alumno = data.get("becas_por_alumno", [])
    becas_filas = [[f"{b['nombre']} ({b['porcentaje']}%)", b["cantidad"]] for b in becas_por_alumno]

    # --- Ganancias por Mes (grafica de barras) ---
    ws["A1"] = "GANANCIAS POR MES"
    ws["A1"].font = SECTION_FONT
    ultima_mes = escribir_tabla(ws, 2, ["Mes", "Ganancias"], ganancias_filas, [24, 16], monedas={1})

    chart_mes = BarChart()
    chart_mes.type = "col"
    chart_mes.title = "Ganancias por Mes"
    chart_mes.style = 10
    chart_mes.y_axis.title = "Monto"
    data_mes = Reference(ws, min_col=2, min_row=2, max_row=ultima_mes)
    cats_mes = Reference(ws, min_col=1, min_row=3, max_row=ultima_mes)
    chart_mes.add_data(data_mes, titles_from_data=True)
    chart_mes.set_categories(cats_mes)
    chart_mes.height = 9
    chart_mes.width = 16
    ws.add_chart(chart_mes, f"D2")

    # --- Estado de Pagos (grafica circular) ---
    fila_est = ultima_mes + 3
    ws.cell(row=fila_est, column=1, value="ESTADO DE PAGOS").font = SECTION_FONT
    ultima_est = escribir_tabla(ws, fila_est + 1, ["Estado", "Cantidad"], estados_filas, [24, 16], enteros={1})

    chart_est = PieChart()
    chart_est.title = "Estado de Pagos"
    data_est = Reference(ws, min_col=2, min_row=fila_est + 1, max_row=ultima_est)
    cats_est = Reference(ws, min_col=1, min_row=fila_est + 2, max_row=ultima_est)
    chart_est.add_data(data_est, titles_from_data=True)
    chart_est.set_categories(cats_est)
    chart_est.height = 8
    chart_est.width = 14
    ws.add_chart(chart_est, f"D{fila_est}")

    # --- Alumnos por Beca (grafica de barras horizontal) ---
    fila_beca = ultima_est + 3
    ws.cell(row=fila_beca, column=1, value="ALUMNOS POR BECA").font = SECTION_FONT
    ultima_beca = escribir_tabla(ws, fila_beca + 1, ["Beca", "Cantidad"], becas_filas, [30, 16], enteros={1})

    chart_beca = BarChart()
    chart_beca.type = "bar"
    chart_beca.title = "Alumnos por Beca"
    chart_beca.style = 11
    data_beca = Reference(ws, min_col=2, min_row=fila_beca + 1, max_row=ultima_beca)
    cats_beca = Reference(ws, min_col=1, min_row=fila_beca + 2, max_row=ultima_beca)
    chart_beca.add_data(data_beca, titles_from_data=True)
    chart_beca.set_categories(cats_beca)
    chart_beca.height = 9
    chart_beca.width = 16
    ws.add_chart(chart_beca, f"D{fila_beca}")


def build_workbook(data, ruta_salida):
    wb = Workbook()

    data["becas_por_alumno"] = [
        {
            "nombre": b["nombre"],
            "porcentaje": b["porcentaje"],
            "cantidad": len([a for a in data["alumnos"] if a.get("beca_id") == b["id"]]),
        }
        for b in data["becas"]
    ]

    crear_hoja_resumen(wb, data)

    crear_hoja_datos(
        wb,
        "Alumnos",
        ["ID", "Nombre", "Primer Apellido", "Segundo Apellido", "Email", "Telefono", "Grado", "Beca (%)", "Fecha"],
        [
            [
                a["id"], a["nombre"], a["primer_apellido"] or "", a["segundo_apellido"] or "",
                a["email"], a["telefono"] or "", a["grado"],
                a["beca_porcentaje"] or 0, formatear_fecha(a["created_at"]),
            ]
            for a in data["alumnos"]
        ],
        [6, 18, 18, 18, 28, 14, 12, 12, 14],
        enteros={0, 7},
        fechas={8},
    )

    crear_hoja_datos(
        wb,
        "Pagos",
        ["ID", "Alumno", "Concepto", "Monto Original", "Beca (%)", "Monto Final", "Semana", "Mes", "Estado", "Fecha"],
        [
            [
                p["id"], p["alumno"], p["concepto"], p["monto_original"] or 0,
                p["beca_porcentaje"] or 0, p["monto_final"] or 0,
                p["semana"] if p["semana"] is not None else 0, p["mes"] or "",
                capitalizar(p["estado"]), formatear_fecha(p["created_at"]),
            ]
            for p in data["pagos"]
        ],
        [6, 30, 30, 16, 10, 16, 8, 12, 14, 14],
        monedas={3, 5},
        enteros={0, 4, 6},
        fechas={9},
    )

    anio = datetime.now().year
    crear_hoja_datos(
        wb,
        "Comprobantes",
        ["ID", "Folio", "Alumno", "Concepto", "Monto", "Metodo de Pago", "Observaciones", "Fecha"],
        [
            [
                c["id"], f"COMP-{anio}-{c['id']:03d}", c["alumno"], c["concepto"],
                c["monto"] or 0, capitalizar(c["metodo_pago"]), c["observaciones"] or "",
                formatear_fecha(c["created_at"]),
            ]
            for c in data["comprobantes"]
        ],
        [6, 16, 30, 30, 16, 18, 30, 14],
        monedas={4},
        enteros={0},
        fechas={7},
    )

    crear_hoja_datos(
        wb,
        "Inscripciones",
        ["ID", "Alumno", "Ciclo Escolar", "Grado", "Monto", "Estado", "Fecha"],
        [
            [
                i["id"], i["alumno"], i["ciclo_escolar"] or "", i["grado"] or "",
                i["monto_inscripcion"] or 0, capitalizar(i["estado"]),
                formatear_fecha(i["fecha_inscripcion"]),
            ]
            for i in data["inscripciones"]
        ],
        [6, 30, 16, 14, 14, 12, 14],
        monedas={4},
        enteros={0},
        fechas={6},
    )

    crear_hoja_datos(
        wb,
        "Becas",
        ["ID", "Nombre", "Porcentaje", "Estado", "Descripcion"],
        [
            [b["id"], b["nombre"], b["porcentaje"] or 0, capitalizar(b["estado"]), b["descripcion"] or ""]
            for b in data["becas"]
        ],
        [6, 22, 12, 12, 40],
        enteros={0, 2},
    )

    crear_hoja_graficas(wb, data)

    wb.save(ruta_salida)
    return ruta_salida


def main():
    parser = argparse.ArgumentParser(description="Reporte general en Excel con graficas")
    parser.add_argument("--salida", default=None, help="Ruta del archivo xlsx de salida")
    args = parser.parse_args()

    conn = conectar()
    data = {clave: _fetch(conn, consulta) for clave, consulta in QUERIES.items()}
    conn.close()

    ruta = args.salida or os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        f"reporte_general_{datetime.now().strftime('%Y-%m-%d')}.xlsx",
    )
    build_workbook(data, ruta)
    print(f"Reporte generado: {ruta}")


if __name__ == "__main__":
    main()
